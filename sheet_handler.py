"""
Schedule handler for the Driver Pickup Reminder Agent.

The driver schedule can live in a local Excel file, a live Google Sheet, or
both. SHEET_BACKEND picks which are watched:

    SHEET_BACKEND=auto    -> watch every source that is configured (default)
    SHEET_BACKEND=excel   -> local .xlsx file only
    SHEET_BACKEND=google  -> live Google Sheet only

"auto" is the default so you never have to edit .env to switch: whichever
schedule you actually edited is the one that gets acted on. A source counts as
configured when its file/credentials are present, so a checkout with no Google
service account simply runs on Excel with no errors.

If the same ride (same phone + pickup time) appears in more than one source,
the driver is called once and the status is written back to every source it
appeared in.

Expected columns (header in row 1, data from row 2):
  1 Driver Name | 2 Driver Phone Number | 3 Pickup Location
  4 Scheduled Pickup Time | 5 Reminder Status
"""

import os
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import config

IST = ZoneInfo(config.TIMEZONE)

# Column positions (1-based), shared by both sources.
COL_DRIVER_NAME = 1
COL_DRIVER_PHONE = 2
COL_PICKUP_LOCATION = 3
COL_PICKUP_TIME = 4
COL_REMINDER_STATUS = 5


# --------------------------- shared helpers ---------------------------

def _parse_pickup_time(value, row_label=""):
    """
    Parse a pickup-time cell into an IST-aware datetime.

    Excel hands us real datetime objects; Google Sheets hands us strings whose
    format depends on how the cell is displayed. Returns None if the value
    cannot be understood.
    """
    if isinstance(value, datetime):
        return value.replace(tzinfo=IST) if value.tzinfo is None else value

    value = str(value).strip()
    if not value:
        return None

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(value, fmt).replace(tzinfo=IST)
        except ValueError:
            continue

    # Tolerant fallback for other display formats (e.g. "18/08/2026 21:39").
    try:
        from dateutil import parser as date_parser
    except ImportError:
        print(f"  [WARN] {row_label}: Cannot parse pickup time '{value}' "
              f"(install python-dateutil for flexible date parsing), skipping")
        return None

    try:
        dt = date_parser.parse(value, dayfirst=True)  # dd/mm/yyyy for India
        return dt.replace(tzinfo=IST) if dt.tzinfo is None else dt
    except (ValueError, OverflowError):
        print(f"  [WARN] {row_label}: Cannot parse pickup time '{value}', skipping")
        return None


def _is_due(pickup_dt, now):
    """
    True if pickup_dt falls inside the reminder window:
    (REMINDER_MINUTES_BEFORE - 5) .. (REMINDER_MINUTES_BEFORE + 5) from now.
    """
    minutes = config.REMINDER_MINUTES_BEFORE
    return (now + timedelta(minutes=minutes - 5)
            <= pickup_dt <=
            now + timedelta(minutes=minutes + 5))


def _collect_due_rows(rows, source, now):
    """
    Turn raw rows (list of cell lists, row 1 = header) into due-ride dicts.

    Shared by both sources so the "Pending" check, the reminder window and the
    field validation can never drift apart between them.
    """
    rides = []

    for row_idx, row in enumerate(rows[1:], start=2):
        # Pad short rows so column indexing never raises.
        cells = list(row) + [None] * (COL_REMINDER_STATUS - len(row))
        name, phone, location, pickup_time, status = cells[:COL_REMINDER_STATUS]

        # Only "Pending" rows are picked up - this is what stops a driver from
        # being called twice for the same ride.
        if status is None or str(status).strip().lower() != "pending":
            continue

        if not all([name, phone, location, pickup_time]):
            print(f"  [WARN] {source} row {row_idx}: Missing required fields, skipping")
            continue

        pickup_dt = _parse_pickup_time(pickup_time, f"{source} row {row_idx}")
        if pickup_dt is None or not _is_due(pickup_dt, now):
            continue

        rides.append({
            "driver_name": str(name).strip(),
            "driver_phone": str(phone).strip(),
            "pickup_location": str(location).strip(),
            "pickup_time": pickup_dt,
            "pickup_time_str": pickup_dt.strftime("%I:%M %p"),  # "09:39 PM"
            "targets": [(source, row_idx)],
        })

    return rides


# --------------------------- Excel source ---------------------------

def _excel_is_configured():
    return os.path.exists(config.EXCEL_FILE_PATH)


def _excel_read_rows():
    from openpyxl import load_workbook

    wb = load_workbook(config.EXCEL_FILE_PATH)
    ws = wb.active
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    wb.close()
    return rows


def _excel_can_write():
    """
    True if the Excel file can be written to right now.

    Excel holds an exclusive lock on an open workbook, so the status
    write-back would fail *after* the call was already placed, leaving the row
    as "Pending" and causing the driver to be called again next cycle.
    """
    path = config.EXCEL_FILE_PATH
    if not os.path.exists(path):
        return False

    # Excel drops a "~$name.xlsx" sidecar next to any workbook it has open.
    directory, name = os.path.split(path)
    if os.path.exists(os.path.join(directory, "~$" + name)):
        return False

    # And the workbook itself cannot be opened for writing while it is locked.
    try:
        with open(path, "r+b"):
            return True
    except OSError:
        return False


def _excel_mark_reminder_status(row_index, status):
    from openpyxl import load_workbook

    wb = load_workbook(config.EXCEL_FILE_PATH)
    ws = wb.active
    ws.cell(row=row_index, column=COL_REMINDER_STATUS, value=status)
    wb.save(config.EXCEL_FILE_PATH)
    wb.close()


# ------------------------ Google Sheets source ------------------------

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# Cache the worksheet handle so we authorize only once per process.
_worksheet = None


def _google_is_configured():
    return bool(config.GOOGLE_SHEET_ID) and os.path.exists(config.GOOGLE_SERVICE_ACCOUNT_JSON)


def _get_worksheet():
    """Authorize (once) and return the target Google worksheet."""
    global _worksheet
    if _worksheet is not None:
        return _worksheet

    import gspread
    from google.oauth2.service_account import Credentials

    creds = Credentials.from_service_account_file(
        config.GOOGLE_SERVICE_ACCOUNT_JSON, scopes=SCOPES
    )
    spreadsheet = gspread.authorize(creds).open_by_key(config.GOOGLE_SHEET_ID)
    _worksheet = (spreadsheet.worksheet(config.GOOGLE_WORKSHEET_NAME)
                  if config.GOOGLE_WORKSHEET_NAME else spreadsheet.sheet1)
    return _worksheet


def _google_read_rows():
    return _get_worksheet().get_all_values()


def _google_mark_reminder_status(row_index, status):
    _get_worksheet().update_cell(row_index, COL_REMINDER_STATUS, status)


# --------------------------- source registry ---------------------------

_SOURCES = {
    "excel": {
        "is_configured": _excel_is_configured,
        "read_rows": _excel_read_rows,
        "can_write": _excel_can_write,
        "mark": _excel_mark_reminder_status,
    },
    "google": {
        "is_configured": _google_is_configured,
        "read_rows": _google_read_rows,
        # Nothing locks a Google Sheet locally; write errors surface per call.
        "can_write": lambda: True,
        "mark": _google_mark_reminder_status,
    },
}


def active_sources():
    """
    Which schedule sources this run should watch.

    In "auto" mode that is every source with its file/credentials in place;
    otherwise it is just the one named by SHEET_BACKEND.
    """
    if config.SHEET_BACKEND == "auto":
        return [n for n, s in _SOURCES.items() if s["is_configured"]()]
    return [config.SHEET_BACKEND]


# ----------------------------- public API -----------------------------

def get_pending_rides():
    """
    Return rides that are "Pending" AND inside the reminder window, across
    every active source.

    Each ride carries a "targets" list of (source, row_index) pairs saying
    where its status must be written back. A ride present in more than one
    source is returned once, with a target for each - so the driver is called
    once and every copy of the schedule is kept in sync.
    """
    now = datetime.now(IST)
    merged = {}

    for name in active_sources():
        source = _SOURCES[name]
        try:
            rows = source["read_rows"]()
        except Exception as e:
            print(f"  [ERROR] Could not read the {name} schedule: {e}")
            continue

        for ride in _collect_due_rows(rows, name, now):
            # Same driver + same pickup time = the same real-world ride.
            key = (ride["driver_phone"], ride["pickup_time"])
            if key in merged:
                merged[key]["targets"].extend(ride["targets"])
            else:
                merged[key] = ride

    return list(merged.values())


def can_write(targets):
    """
    Whether the status can be persisted to every target of a ride.

    The agent checks this before dialling: if the outcome cannot be recorded,
    placing the call would only get the driver called again on every later
    cycle.
    """
    return all(_SOURCES[name]["can_write"]() for name, _ in targets)


def mark_reminder_status(targets, status):
    """
    Write the reminder status back to each (source, row_index) target.

    Args:
        targets (list): (source, row_index) pairs from a ride dict
        status (str): "Sent", "Failed - No Answer", "Failed - Busy",
                      or "Failed - Error"

    Raises:
        RuntimeError: if any target could not be updated. The agent reports
            this rather than letting it kill the cycle.
    """
    failures = []

    for name, row_index in targets:
        try:
            _SOURCES[name]["mark"](row_index, status)
        except Exception as e:
            failures.append(f"{name} row {row_index}: {e}")

    if failures:
        raise RuntimeError("; ".join(failures))
